import xlsxwriter,os, json
from openpyxl import load_workbook


def create_output_file(testcase,outputfolder,output):
    os.chdir(outputfolder)
    outputfile = open(str(testcase)+'.json',"w+")
    json.dump(output, outputfile)

def create_test_report(i,testcase,testcasedecription,status,output_file_path):
    wb = load_workbook(output_file_path)
    sheet = wb.active
    sheet.cell(row=i,column=1).value=testcase
    sheet.cell(row=i,column=2).value=testcasedecription
    sheet.cell(row=i,column=3).value=status
    wb.save(output_file_path)
    wb.close()
