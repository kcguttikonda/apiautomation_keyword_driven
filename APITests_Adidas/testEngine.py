import collections
import os, time
from openpyxl import Workbook

from APITests_Adidas.getTestCases import testdatabuilder
from APITests_Adidas.invokeAPI import invokeapi
from APITests_Adidas.reporting import create_output_file, create_test_report

testrun_output = collections.OrderedDict()
tests_to_run = testdatabuilder()
apiendpoint = "http://api.openweathermap.org/data/2.5/weather?"
appid = "df75744c36761e9ee5797ed2c5261a2b"
output_folder = os.path.dirname(os.getcwd()) + "\\Outputs\\" + str(time.time())
os.mkdir(output_folder)
os.chdir(output_folder)
output_file_path = output_folder+"\\" + "TestExecutionReport.xlsx"
wb = Workbook()
sheet = wb.active
sheet.cell(row=1,column=1).value="TestCaseID"
sheet.cell(row=1,column=2).value="TestCaseDescription"
sheet.cell(row=1,column=3).value="Status"
wb.save(output_file_path)
i=2


for testcase in tests_to_run.keys():
    apiurl = apiendpoint + tests_to_run[testcase][6] + "&appid=" + appid
    # invokeapi(method, uri, headers, body, payload):
    response = invokeapi(tests_to_run[testcase][2],
                         apiurl,
                         tests_to_run[testcase][5],
                         tests_to_run[testcase][4],
                         tests_to_run[testcase][6])
    status_code = response.status_code
    response_content = response.json()


    if status_code == 200 or status_code == 201:

        testrun_output[testcase] = "PASS"
    else:
        testrun_output[testcase] = "FAIL"

    create_output_file(testcase, output_folder, response_content)
    create_test_report(i,testcase, tests_to_run[testcase][1], testrun_output[testcase],output_file_path)
    i+=1
print(testrun_output)
